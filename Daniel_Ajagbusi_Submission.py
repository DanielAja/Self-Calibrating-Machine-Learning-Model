#Wells Fargo Campus Analytics Challenge 2020
#Submission by: Daniel Ajagbusi
#Date: 7/29/2020
#Institution: University of Minnesota - Twin Cities

#KEY ~~~~~~~~~~~~~~~~~~~~~
# * : This symbol denotes a while loop in use to iterate over a list with a finite
# number of values that should have been able to be predetermined, this was devised
# as a fix to errors found in testing that prevented consistent results.
#
#
#

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from sklearn.linear_model import LogisticRegression
from sklearn.feature_selection import GenericUnivariateSelect, f_regression
from sklearn.metrics import f1_score

#Sets and lists
omitted_data_col = set() #Data deemed inconsequential to determining the appropriate outcome (Collums) 

training_data_set = [] 
training_result_set= []
evaluation_data_set = []
evaluation_result_set = []

#Given variables
current_model = LogisticRegression(solver='lbfgs')

# More information about how these values were determined can be found in "My Method.pdf"
feature_variables_percentage_value = 50
min_correlated_value_value = 5.64

training_data_set_file = 'ChallengeDataSet.xlsx'
evaluation_data_set_file = 'EvaluationDataSet.xlsx'
results_output_file = "ResultsOutput.xlsx"

def read_data():
    # This function iterates over the data for the training data set storing its
    # values in the global training_data_set and training_result_set lists accordingly.
    # This function encodes all chars as there ASCII equivalent numeric value minus 64, 
    # (A = 1, B=2, ...) and assumes the final column is the results.

    global training_data_set
    global training_result_set
    
    training_data_set = []
    training_result_set= []

    training_WB = load_workbook(filename=training_data_set_file)
    training_sheet = training_WB.active
    training_sheet_col = training_sheet.max_column
    
    cellvalue = 1
    i = 1
    # *
    while cellvalue != None: 
        current_row_data = []
        i = i +1
        for j in range(1, training_sheet_col):
            cellvalue = training_sheet.cell(row = i, column = j).value
            if type(cellvalue) == str:
                cellvalue = ord(cellvalue) - 64
            if cellvalue != None and j not in omitted_data_col:
                current_row_data.append(cellvalue)
        
        cellvalue = training_sheet.cell(row = i, column = training_sheet_col).value
        if cellvalue != None:
            training_data_set.append(current_row_data)
            training_result_set.append(cellvalue)

def train_model(data_set, result_set):
    # This function uses LogisticRegression() to train a global machine 
    # learning model, based on the data_set and result_set provided

    global current_model

    model_in_use = LogisticRegression(solver='lbfgs')
    model_in_use.fit(data_set, result_set)
    current_model = model_in_use

def generate_data(file_name):
    # This function uses the global machine learning model to evaluate new data passed to it
    # in the form of a .xlsx file, storing the data and results in the global evaluation_data_set,
    # and evaluation_result_set lists respectfully

    global evaluation_result_set
    global evaluation_data_set

    evaluation_data_set = []
    evaluation_result_set = []
    
    evaluation_WB = load_workbook(filename=file_name)
    evaluation_sheet = evaluation_WB.active
    evaluation_row_size = evaluation_sheet.max_row
    evaluation_col_size = evaluation_sheet.max_column

    cellvalue = 1
    i = 1

    # *
    while cellvalue != None: 
        current_row_data = []
        i = i +1
        for j in range(1, evaluation_col_size):
            cellvalue = evaluation_sheet.cell(row = i, column = j).value
            if type(cellvalue) == str:
                cellvalue = ord(cellvalue) - 64
            if cellvalue != None and j not in omitted_data_col:
                current_row_data.append(cellvalue)
        
        if cellvalue != None:
            evaluation_data_set.append(current_row_data)
    
    evaluation_result_set = current_model.predict(evaluation_data_set)

def output_results(input_file_name, output_file_name):
    # This function uses the global evaluation_result_set list to create a .xlsx file formatted as specified
    # in the Official Rules (Maintaining the same row order and prefaced with the titles of "dataset_id" and 
    # "prediction_score" respectfully) 

    evaluation_WB = load_workbook(filename=input_file_name)
    evaluation_sheet = evaluation_WB.active

    results_WB = Workbook()
    results_sheet = results_WB.active
    results_sheet["A1"] = "dataset_id"
    results_sheet["B1"] = "prediction_score"

    for i in range (2, evaluation_sheet.max_row+1):
        a_cell = results_sheet.cell(row = i, column = 1)
        a_cell.value = i-1
        b_cell = results_sheet.cell(row = i, column = 2)
        b_cell.value = evaluation_result_set[i-2]

    results_WB.save(filename=output_file_name)

def xlsx_to_csv(file_name):
    # This function creates a copy of a .xlsx file as a .csv 
    # when passed a string of the file name ending with .xlsx

    xlsx = load_workbook(filename=file_name)
    sheet = xlsx.active
    data = sheet.rows
    csv = open(file_name.replace(".xlsx" ,".csv"), "w+")

    for row in data:
        l = list(row)
        for i in range(0, len(l)):
            if i == len(l)-1:
                csv.write(str(l[i].value))
                csv.write('\n')
            else:
                csv.write(str(l[i].value) + ',')
            
    csv.close()

def run(bool_output_file):
    # After specifying the names of the .xlsx files as the global variables of 
    # training_data_set_file, evaluation_data_set_file, and results_output_file at 
    # the top of the code, this function allows you to specify an integer percentage  
    # (feature_variables_percentage_value) that denotes the percentage of the top variable, and minimal 
    # correlation value (min_correlated_value_value) to denote the smallest correlation you would like 
    # to include in the training of the machine learning model, then outputs a file in the format 
    # specified by the Official Rules as .xlsx and .csv (Maintaining the same row order and prefaced with the 
    # titles of "dataset_id" and "prediction_score" respectfully) Additionally, you may pass a 
    # boolean variable to denote if you would like it to generate a file, this feature was added 
    # to this function so it could be used in other places in the program

    global training_data_set
    global training_result_set
    global evaluation_data_set_file
    global results_output_file
    global omitted_data_col
    global evaluation_data_set
    global evaluation_result_set 
    global feature_variables_percentage_value
    global min_correlated_value_value

    training_data_set = [] 
    training_result_set= []
    evaluation_data_set = []
    evaluation_result_set = []
    
    omitted_data_col.clear()

    read_data()

    min_correlated_run_value = min_correlated_value_value/100
    data_set = pd.DataFrame(training_data_set)
    correlation_matrix = data_set.corr()

    for j in range(len(correlation_matrix.columns)):
        for k in range(j):
            if abs(correlation_matrix.iloc[j, k]) > min_correlated_run_value:
                colname = correlation_matrix.columns[j]
                omitted_data_col.add(colname)

    transformer = GenericUnivariateSelect(f_regression, mode='percentile', param=feature_variables_percentage_value)
    transformed_training_data_set = transformer.fit_transform(training_data_set, training_result_set)
    
    include_data = transformer.get_support(False)
    for j in range(0,len(include_data)):
        if include_data[j] == False:
            omitted_data_col.add(j+1)

    transformed_training_data_set = []
    for j in range(0, len(training_data_set)):
            current_row_data = []
            for k in range(0, len(training_data_set[0])):
                if k not in omitted_data_col:
                    current_row_data.append(training_data_set[j][k])
            transformed_training_data_set.append(current_row_data)

    train_model(transformed_training_data_set, training_result_set)
    generate_data(evaluation_data_set_file)
    if bool_output_file == True:
        output_results(evaluation_data_set_file, results_output_file)
        xlsx_to_csv(results_output_file)

def find_f1_score():
    # This function uses the data from the training_data_set_file to compare the expected outcome,
    # that of which is provided in the file, to the outcome generated when under the restrictions 
    # specified by feature_variables_percentage_value and min_correlated_value_value variables. 
    # This function is able to run with no user input assuming the variables at the top of the page 
    # have already been set to there proper values. 
    
    global training_data_set
    global training_result_set 
    global evaluation_data_set
    global evaluation_result_set
    global omitted_data_col

    reduced_training_data_set = []

    run(bool_output_file = False)

    for i in range (0,len(training_data_set)):
        current_row = []
        for j in range(0, len(training_data_set[0])):
            if j not in omitted_data_col:
                current_row.append(training_data_set[i][j])
        reduced_training_data_set.append(current_row)
    
    prediction_of_training_set_results = current_model.predict(reduced_training_data_set)
    f1_score_num = f1_score(training_result_set, prediction_of_training_set_results)

    print("With a F1 score of: " + str(f1_score_num))

#~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~~
# The premise of this section was to create systems to determine the minimum 
# feature variable percentage and correlation values that are still able to 
# result in an accurate machine learning model. Underneath each function is
# a commented out example of how it can be used.

def calibrate_feature_selection(precision,result_accuracy_cut_off):
    # This function takes 2 parameters, precision - which acts as the amount the value 
    # is incremented while calibrating, and result_accuracy_cut_off - which is the smallest 
    # percentage of accuracy a given machine learning model has when compared with the data 
    # provided in the training_data_set_file. 
    #
    # This function allows for the feature_variables_percentage_value to be set to the user-specified
    # accuracy percentage, actively feature selecting. resulting in a smaller size for the data used
    # to build the machine learning model, due to the relative feature selections. 


    print("~~~ Beginning Feature Variable Percentage Calibration ~~~")
    global training_data_set
    global omitted_data_col
    
    training_data_set = []

    evaluation_WB = load_workbook(filename=training_data_set_file)
    evaluation_sheet = evaluation_WB.active

    results_WB = Workbook()
    results_sheet = results_WB.active

    read_data()

    i = precision
    match_count = 0
    found = False
    while i < 100:
        omitted_data_col = set()
        transformer = GenericUnivariateSelect(f_regression, mode='percentile', param=i)
        transformed_training_data_set = transformer.fit_transform(training_data_set, training_result_set)

        train_model(transformed_training_data_set,training_result_set)
        include_data = transformer.get_support(False)
        for j in range(0,len(include_data)):
            if include_data[j] == False:
                omitted_data_col.add(j+1)
        
        generate_data(training_data_set_file)

        for j in range (0, len(evaluation_result_set)):
            if evaluation_result_set[j] == training_result_set[j]:
                match_count = match_count +1

        if (match_count/len(evaluation_result_set))*100 >= result_accuracy_cut_off:
            feature_variables_percentage_value = i
            print ("Feature Variable Percentage Final Value: " + str(feature_variables_percentage_value) + "%, With a " + str(round((match_count/len(evaluation_result_set))*100,2)) + "% " + "accuracy")
            found = True
            break

        print(str(i) + "% " + "of avaible data: " + str(round((match_count/len(evaluation_result_set))*100,5)) + "% " + "accurate results")
        i = i + precision
        match_count = 0
    if found == False:
        print("~ Unable to find Feature Variable Percentage that fits your specifications, please try lowering the cut off value.")
        feature_variables_percentage_value = 100
        print("Feature Variable Percentage has been set to include all data")
calibrate_feature_selection(precision = 0.1, result_accuracy_cut_off = 99)

def calibrate_correlation_selection(precision,result_accuracy_cut_off):
    # This function takes 2 parameters, precision - which acts as the amount the value is incremented 
    # while calibrating, and result_accuracy_cut_off - which is the smallest percentage of accuracy a 
    # given machine learning model has when compared with the data provided in the training_data_set_file. 
    #
    # This function allows for the min_correlated_value_value to be set to the user-specified accuracy 
    # percentage, actively removing correlated variables. resulting in a smaller size for the data used 
    # to build the machine learning model. 

    
    print("~~~ Beginning Correlation Value Calibration ~~~")
    global omitted_data_col
    global training_data_set

    training_data_set = []

    precision = precision/100

    evaluation_WB = load_workbook(filename=training_data_set_file)
    evaluation_sheet = evaluation_WB.active

    results_WB = Workbook()
    results_sheet = results_WB.active

    read_data()

    i = precision
    match_count = 0
    found = False
    while i < 100:
        omitted_data_col = set()

        data_set = pd.DataFrame(training_data_set)
        correlation_matrix = data_set.corr()

        for j in range(len(correlation_matrix.columns)):
            for k in range(j):
                if abs(correlation_matrix.iloc[j, k]) > i:
                    colname = correlation_matrix.columns[j]
                    omitted_data_col.add(colname)
        
        transformed_training_data_set = []

        for j in range(0, len(training_data_set)):
            current_row_data = []
            for k in range(0, len(training_data_set[0])):
                if k not in omitted_data_col:
                    current_row_data.append(training_data_set[j][k])
            transformed_training_data_set.append(current_row_data)

        train_model(transformed_training_data_set,training_result_set)
        generate_data(training_data_set_file)

        for j in range (0, len(evaluation_result_set)):
            if evaluation_result_set[j] == training_result_set[j]:
                match_count = match_count +1

        if (match_count/len(evaluation_result_set))*100 >= result_accuracy_cut_off:
            min_correlated_value_value = i*100
            print ("Correlation Value Final Value: " + str(round(min_correlated_value_value,2)) + ", With a " + str(round((match_count/len(evaluation_result_set))*100,2)) + "% " + "accuracy")
            found = True
            break

        print("Correlation Value of " + str(round(i*100,5)) + ": " + str(round((match_count/len(evaluation_result_set))*100,5)) + "% " + "accurate results")
        i = i + precision
        match_count = 0

    if found == False:
        print("~ Unable to find a Correlation Value that fits your specifications, please try lowering the cut off value.")
        min_correlated_value_value = 100
        print("Correlation Value has been set to include all data")   
#calibrate_correlation_selection(precision = .01,result_accuracy_cut_off = 99)
 
def record_learning_curve(file_name,iterations):
    # This function trains machine learning models using a set percentage of the most imporant
    # available data then creates a .xlsx file of the resulting outputs, when crossed
    # compared with the data that was used to train it, to allow for the quantitative
    # analysis of the "Learning Curve" data 
    
    global training_data_set
    global omitted_data_col
    
    training_data_set = []

    increment = 100/iterations
    
    evaluation_WB = load_workbook(filename=training_data_set_file)
    evaluation_sheet = evaluation_WB.active

    results_WB = Workbook()
    results_sheet = results_WB.active

    for i in range (0, iterations):
        
        percentage = increment + i*increment
        omitted_data_col = set()
        
        read_data()

        transformer = GenericUnivariateSelect(f_regression, mode='percentile', param=percentage)
        transformed_training_data_set = transformer.fit_transform(training_data_set, training_result_set)
        
        print(str(np.array(transformed_training_data_set).shape) + " @ " + str(percentage) + "%")
        
        train_model(transformed_training_data_set,training_result_set)
        include_data = transformer.get_support(False)
        for j in range(0,len(include_data)):
            if include_data[j] == False:
                omitted_data_col.add(j+1)
        
        generate_data(training_data_set_file)

        cell = results_sheet.cell(row = 1, column = i+1)
        cell.value = percentage
        cell = results_sheet.cell(row = 2, column = i+1)
        cell.value = ("=SUMPRODUCT(--(" + chr(i+65) + "3:" + chr(i+65) + str(len(evaluation_result_set)+2) + "=$" + chr(iterations+64) + "$3:$" + chr(iterations+64) + "$" + str(len(evaluation_result_set)+2) + "))/" +str(len(evaluation_result_set)))

        for j in range (1, len(evaluation_result_set)+1):
            cell = results_sheet.cell(row = j+2, column = i+1)
            cell.value = evaluation_result_set[j-2]
    results_WB.save(filename=file_name)
#record_learning_curve('LearningCurve.xlsx',20)

def record_correlation(file_name,iterations,min_percentage,max_percentage):
    # This function trains machine learning models only using data that falls out of
    # range of a certain correlation value then creates a .xlsx file of the resulting outputs,
    # when crossed compared with the data that was used to train it, to allow for the quantitative
    # analysis of the "Correlation Curve" of the data 

    global omitted_data_col
    global training_data_set
    training_data_set = []

    min_percentage = min_percentage/100
    max_percentage = max_percentage/100
    increment = (max_percentage - min_percentage)/iterations

    evaluation_WB = load_workbook(filename=training_data_set_file)
    evaluation_sheet = evaluation_WB.active

    results_WB = Workbook()
    results_sheet = results_WB.active
    
    for i in range (0, iterations):
        percentage = min_percentage + i*increment
        omitted_data_col = set()

        read_data()

        data_set = pd.DataFrame(training_data_set)
        correlation_matrix = data_set.corr()

        for j in range(len(correlation_matrix.columns)):
            for k in range(j):
                if abs(correlation_matrix.iloc[j, k]) > percentage:
                    colname = correlation_matrix.columns[j]
                    omitted_data_col.add(colname)
        
        transformed_training_data_set = []

        for j in range(0, len(training_data_set)):
            current_row_data = []
            for k in range(0, len(training_data_set[0])):
                if k not in omitted_data_col:
                    current_row_data.append(training_data_set[j][k])
            transformed_training_data_set.append(current_row_data)
        
        print (str(np.array(transformed_training_data_set).shape) + " @ " + str(round(percentage*100,2)) +"%")
        train_model(transformed_training_data_set,training_result_set)

        generate_data(training_data_set_file)
        cell = results_sheet.cell(row = 1, column = i+1)
        cell.value = percentage
        cell = results_sheet.cell(row = 2, column = i+1)
        cell.value = ("=SUMPRODUCT(--(" + chr(i+65) + "3:" + chr(i+65) + str(len(evaluation_result_set)+2) + "=$" + chr(iterations+64) + "$3:$" + chr(iterations+64) + "$" + str(len(evaluation_result_set)+2) + "))/" +str(len(evaluation_result_set)))

        for j in range (1, len(evaluation_result_set)+1):
            cell = results_sheet.cell(row = j+2, column = i+1)
            cell.value = evaluation_result_set[j-2]
    results_WB.save(filename=file_name)
#record_correlation("CorrelationCurve.xlsx", 20, 0, 7)

# The calibrations are commented out to save time, as I have already run this 
# and specified the variable values of feature_variables_percentage_value = 50 
# and min_correlated_value_value = 5.64 respectfully. 

#calibrate_feature_selection(precision = 0.1, result_accuracy_cut_off = 99)
#calibrate_correlation_selection(precision = .01,result_accuracy_cut_off = 99)
#run(bool_output_file = True)
#find_f1_score()

print('~~~ Fin. ~~~')  
